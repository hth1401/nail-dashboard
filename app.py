import streamlit as st
import pandas as pd
import numpy as np
import datetime as dt
import os
import openpyxl

st.set_page_config(page_title="Ly Lovely Nails - Dashboard", layout="wide")

DATA_PATH = "data.xlsx"
STORE_CSV = "input_data.csv"   # app sẽ lưu dữ liệu nhập ở đây (an toàn hơn sửa thẳng Excel)
EXPORT_XLSX = "export.xlsx"    # file Excel xuất ra cho kế toán / lưu trữ

# ---------- Helpers ----------
def load_setup(xlsx_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["SETUP"]
    # tìm dòng header "Tech (NAME)"
    header_row = None
    for r in range(1, 50):
        if (ws.cell(r,1).value or "").strip() == "Tech (NAME)":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Không tìm thấy bảng SETUP trong sheet SETUP.")
    rows=[]
    for r in range(header_row+1, header_row+500):
        tech = ws.cell(r,1).value
        code = ws.cell(r,2).value
        tech_pct = ws.cell(r,3).value
        salon_pct = ws.cell(r,4).value
        if tech is None or str(tech).strip()=="":
            break
        rows.append({
            "TECH": str(tech).strip().upper(),
            "CODE": str(code).strip().upper() if code is not None else "",
            "TECH_PCT": float(tech_pct) if tech_pct is not None else 0.0,
            "SALON_PCT": float(salon_pct) if salon_pct is not None else 0.0,
        })
    df=pd.DataFrame(rows)
    if df.empty:
        raise ValueError("Bảng SETUP đang trống.")
    return df

def normalize_tech(raw: str, setup: pd.DataFrame) -> str:
    if raw is None:
        return ""
    s=str(raw).strip().upper()
    if s=="":
        return ""
    # cho phép nhập kiểu "RY - RYAN"
    if " - " in s:
        s=s.split(" - ",1)[1].strip().upper()
    # map theo CODE
    m = setup.set_index("CODE")["TECH"].to_dict()
    if s in m:
        return m[s]
    # nếu gõ full name
    techs=set(setup["TECH"].tolist())
    if s in techs:
        return s
    return s

def monday_of(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())

def month_key(d: dt.date) -> str:
    return f"{d.year:04d}-{d.month:02d}"

def coerce_date(x):
    if pd.isna(x) or x is None or x=="":
        return pd.NaT
    if isinstance(x, (dt.date, dt.datetime)):
        return pd.to_datetime(x).date()
    return pd.to_datetime(x, dayfirst=True, errors="coerce").date()

def load_input_from_template(xlsx_path: str) -> pd.DataFrame:
    # Sheet INPUT: header ở dòng có "Date"
    wb=openpyxl.load_workbook(xlsx_path, data_only=True)
    ws=wb["INPUT"]
    header_row=None
    for r in range(1, 30):
        if (ws.cell(r,1).value or "").strip()=="Date":
            header_row=r
            break
    if header_row is None:
        raise ValueError("Không tìm thấy header trong sheet INPUT.")
    cols=[]
    for c in range(1, 30):
        v=ws.cell(header_row, c).value
        if v is None:
            break
        cols.append(str(v).strip())
    data=[]
    for r in range(header_row+1, header_row+5000):
        row=[]
        empty=True
        for c in range(1, len(cols)+1):
            v=ws.cell(r,c).value
            row.append(v)
            if v not in (None,""):
                empty=False
        if empty:
            continue
        data.append(row)
    df=pd.DataFrame(data, columns=cols)
    return df

def load_input(setup: pd.DataFrame) -> pd.DataFrame:
    # ưu tiên load từ CSV (dữ liệu mới), nếu chưa có thì lấy từ Excel template
    if os.path.exists(STORE_CSV):
        df=pd.read_csv(STORE_CSV)
    else:
        df=load_input_from_template(DATA_PATH)
    # chuẩn hoá cột
    want = ["Date","Tech","Service","Tip","Discount %","Gift Card"]
    for c in want:
        if c not in df.columns:
            df[c]=np.nan
    df=df[want].copy()
    df["Date"] = df["Date"].apply(coerce_date)
    df["Tech"] = df["Tech"].astype(str).replace({"nan":""})
    # số
    for c in ["Service","Tip","Discount %","Gift Card"]:
        df[c]=pd.to_numeric(df[c], errors="coerce")
    df["Discount %"]=df["Discount %"].fillna(0.0)
    df["Gift Card"]=df["Gift Card"].fillna(0.0)
    df["Tip"]=df["Tip"].fillna(0.0)
    df["Service"]=df["Service"].fillna(0.0)
    # tính toán
    df["TechNorm"]=df["Tech"].apply(lambda x: normalize_tech(x, setup))
    pct = setup.set_index("TECH")[["TECH_PCT","SALON_PCT"]].to_dict("index")
    df["Tech %"]=df["TechNorm"].apply(lambda t: pct.get(t, {"TECH_PCT":0.0})["TECH_PCT"])
    df["Salon %"]=df["TechNorm"].apply(lambda t: pct.get(t, {"SALON_PCT":1.0})["SALON_PCT"])
    df["Tech Share"]=df["Service"]*df["Tech %"]
    df["Salon Share (gross)"]=df["Service"]*df["Salon %"]
    df["Salon Net"]=df["Salon Share (gross)"] - (df["Service"]*df["Discount %"]) - df["Gift Card"]
    df["WeekStart(Mon)"]=df["Date"].apply(lambda d: monday_of(d) if isinstance(d, dt.date) else pd.NaT)
    df["Month"]=df["Date"].apply(lambda d: month_key(d) if isinstance(d, dt.date) else "")
    return df

def save_input(df_raw: pd.DataFrame):
    # chỉ lưu những cột nhập (để nhẹ + tránh lỗi)
    keep=["Date","Tech","Service","Tip","Discount %","Gift Card"]
    out=df_raw[keep].copy()
    # Date -> yyyy-mm-dd
    out["Date"]=out["Date"].apply(lambda x: x if isinstance(x,str) else (x.isoformat() if isinstance(x, dt.date) else ""))
    out.to_csv(STORE_CSV, index=False)

def export_excel(df: pd.DataFrame, setup: pd.DataFrame):
    # Xuất 3 sheet: INPUT, DAILY, WEEKLY, MONTHLY
    with pd.ExcelWriter(EXPORT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="INPUT_CLEAN")
        # daily totals per tech
        daily = df.dropna(subset=["Date"]).groupby(["Date","TechNorm"], as_index=False).agg(
            customers=("Service","size"),
            service=("Service","sum"),
            tech_share=("Tech Share","sum"),
            tip=("Tip","sum"),
            salon_net=("Salon Net","sum")
        )
        daily.to_excel(writer, index=False, sheet_name="DAILY_TOTALS")
        weekly = df.dropna(subset=["WeekStart(Mon)"]).groupby(["WeekStart(Mon)","TechNorm"], as_index=False).agg(
            customers=("Service","size"),
            service=("Service","sum"),
            tech_share=("Tech Share","sum"),
            tip=("Tip","sum"),
            salon_net=("Salon Net","sum")
        )
        weekly.to_excel(writer, index=False, sheet_name="WEEKLY_TOTALS")
        monthly = df[df["Month"]!=""].groupby(["Month","TechNorm"], as_index=False).agg(
            customers=("Service","size"),
            service=("Service","sum"),
            tech_share=("Tech Share","sum"),
            tip=("Tip","sum"),
            salon_net=("Salon Net","sum")
        )
        monthly.to_excel(writer, index=False, sheet_name="MONTHLY_TOTALS")

# ---------- UI ----------
st.title("📊 Nail Salon Dashboard (mobile xem nhanh • PC nhập dễ)")

with st.sidebar:
    st.header("Cài đặt")
    st.caption("Dữ liệu nhập sẽ được lưu vào file CSV (an toàn, ít lỗi).")
    if st.button("🔄 Reload dữ liệu"):
        st.rerun()

setup = load_setup(DATA_PATH)
df = load_input(setup)

tab_mobile, tab_input, tab_daily, tab_week, tab_month, tab_export = st.tabs(
    ["📱 Mobile view", "✍️ Nhập dữ liệu (PC)", "🗓️ Daily", "📅 Week", "🗓️ Month", "⬇️ Export"]
)

# ---------- MOBILE VIEW ----------
with tab_mobile:
    st.subheader("Xem nhanh (dành cho điện thoại)")
    cols = st.columns([1,1,1])
    # chọn ngày / thợ / tuần / tháng
    valid_dates = sorted([d for d in df["Date"].dropna().unique()])
    default_date = valid_dates[-1] if valid_dates else dt.date.today()
    pick_date = cols[0].date_input("Chọn ngày", value=default_date)

    techs = ["ALL"] + sorted(set([t for t in df["TechNorm"].unique() if t]))
    pick_tech = cols[1].selectbox("Chọn thợ", techs, index=0)

    weeks = sorted([d for d in df["WeekStart(Mon)"].dropna().unique()])
    default_week = weeks[-1] if weeks else monday_of(dt.date.today())
    pick_week = cols[2].date_input("Chọn tuần (Thứ 2)", value=default_week)

    # Cards nhanh
    def filter_df(dfx: pd.DataFrame):
        out=dfx.copy()
        if pick_tech!="ALL":
            out=out[out["TechNorm"]==pick_tech]
        return out

    ddf = filter_df(df[df["Date"]==pick_date])
    wdf = filter_df(df[df["WeekStart(Mon)"]==pick_week])

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Khách (ngày)", int(ddf.shape[0]))
    c2.metric("Service (ngày)", f"${ddf['Service'].sum():,.2f}")
    c3.metric("Tip (ngày)", f"${ddf['Tip'].sum():,.2f}")
    c4.metric("Salon net (ngày)", f"${ddf['Salon Net'].sum():,.2f}")

    st.divider()
    st.write("**Top theo thợ (ngày)**")
    daily_tech = ddf.groupby("TechNorm", as_index=False).agg(
        customers=("Service","size"),
        service=("Service","sum"),
        tech_share=("Tech Share","sum"),
        tip=("Tip","sum"),
        salon_net=("Salon Net","sum")
    ).sort_values("service", ascending=False)
    st.dataframe(daily_tech, use_container_width=True, hide_index=True)

    st.divider()
    st.write("**Tổng tuần**")
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Khách (tuần)", int(wdf.shape[0]))
    c2.metric("Service (tuần)", f"${wdf['Service'].sum():,.2f}")
    c3.metric("Tip (tuần)", f"${wdf['Tip'].sum():,.2f}")
    c4.metric("Salon net (tuần)", f"${wdf['Salon Net'].sum():,.2f}")

# ---------- INPUT ----------
with tab_input:
    st.subheader("Nhập dữ liệu (1 khách = 1 dòng)")
    st.caption("Gợi ý: nhập Tech có thể gõ CODE (RY/HA/...) hoặc tên đầy đủ. Date chọn bằng lịch.")
    # bảng nhập: chỉ cột cần nhập
    edit_df = df[["Date","Tech","Service","Tip","Discount %","Gift Card"]].copy()
    # chỉnh kiểu
    edited = st.data_editor(
        edit_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Date": st.column_config.DateColumn(format="DD/MM/YYYY"),
            "Service": st.column_config.NumberColumn(step=1),
            "Tip": st.column_config.NumberColumn(step=1),
            "Discount %": st.column_config.NumberColumn(help="0.1 = 10%", step=0.01),
            "Gift Card": st.column_config.NumberColumn(step=1),
        },
        key="editor"
    )
    c1,c2 = st.columns([1,2])
    if c1.button("💾 Save"):
        save_input(edited)
        st.success("Đã lưu dữ liệu!")
        st.rerun()
    c2.info("Bạn nhập trên máy tính, xong bấm **Save**. Điện thoại chỉ vào tab Mobile để xem nhanh.")

# recompute df after editing? We reload via Save, so fine.

# ---------- DAILY ----------
with tab_daily:
    st.subheader("Daily (chi tiết + tổng theo thợ)")
    pick = st.date_input("Chọn ngày", value=default_date, key="daily_date")
    d = df[df["Date"]==pick].copy()
    st.write("**Chi tiết từng khách**")
    st.dataframe(d[["Date","TechNorm","Service","Discount %","Gift Card","Tip","Tech Share","Salon Share (gross)","Salon Net"]],
                 use_container_width=True, hide_index=True)
    st.write("**Tổng theo thợ**")
    tot = d.groupby("TechNorm", as_index=False).agg(
        customers=("Service","size"),
        service=("Service","sum"),
        tech_share=("Tech Share","sum"),
        tip=("Tip","sum"),
        salon_net=("Salon Net","sum")
    )
    st.dataframe(tot, use_container_width=True, hide_index=True)

# ---------- WEEK ----------
with tab_week:
    st.subheader("Week summary (1 ngày tổng / hoặc theo thợ)")
    pickw = st.date_input("Chọn WeekStart (Thứ 2)", value=default_week, key="week_start")
    tech_pick = st.selectbox("Lọc thợ", ["ALL"]+sorted(set([t for t in df["TechNorm"].unique() if t])), key="week_tech")
    w = df[df["WeekStart(Mon)"]==pickw].copy()
    if tech_pick!="ALL":
        w = w[w["TechNorm"]==tech_pick]
    by_day = w.groupby("Date", as_index=False).agg(
        customers=("Service","size"),
        service=("Service","sum"),
        tech_share=("Tech Share","sum"),
        tip=("Tip","sum"),
        salon_net=("Salon Net","sum")
    ).sort_values("Date")
    st.write("**Tổng theo ngày trong tuần**")
    st.dataframe(by_day, use_container_width=True, hide_index=True)
    st.write("**Tổng theo thợ trong tuần**")
    by_tech = w.groupby("TechNorm", as_index=False).agg(
        customers=("Service","size"),
        service=("Service","sum"),
        tech_share=("Tech Share","sum"),
        tip=("Tip","sum"),
        salon_net=("Salon Net","sum")
    ).sort_values("service", ascending=False)
    st.dataframe(by_tech, use_container_width=True, hide_index=True)

# ---------- MONTH ----------
with tab_month:
    st.subheader("Month summary")
    months = sorted([m for m in df["Month"].unique() if m])
    default_m = months[-1] if months else f"{dt.date.today().year:04d}-{dt.date.today().month:02d}"
    pickm = st.selectbox("Chọn tháng (yyyy-mm)", months if months else [default_m], index=(len(months)-1 if months else 0))
    tech_pick = st.selectbox("Lọc thợ", ["ALL"]+sorted(set([t for t in df["TechNorm"].unique() if t])), key="month_tech")
    mdf = df[df["Month"]==pickm].copy()
    if tech_pick!="ALL":
        mdf = mdf[mdf["TechNorm"]==tech_pick]
    by_day = mdf.groupby("Date", as_index=False).agg(
        customers=("Service","size"),
        service=("Service","sum"),
        tech_share=("Tech Share","sum"),
        tip=("Tip","sum"),
        salon_net=("Salon Net","sum")
    ).sort_values("Date")
    st.write("**Tổng theo ngày trong tháng**")
    st.dataframe(by_day, use_container_width=True, hide_index=True)
    st.write("**Tổng theo thợ trong tháng**")
    by_tech = mdf.groupby("TechNorm", as_index=False).agg(
        customers=("Service","size"),
        service=("Service","sum"),
        tech_share=("Tech Share","sum"),
        tip=("Tip","sum"),
        salon_net=("Salon Net","sum")
    ).sort_values("service", ascending=False)
    st.dataframe(by_tech, use_container_width=True, hide_index=True)

# ---------- EXPORT ----------
with tab_export:
    st.subheader("Xuất file")
    st.write("Bấm nút để xuất ra Excel mới (không phụ thuộc formula nên ít lỗi).")
    if st.button("⬇️ Tạo export.xlsx"):
        export_excel(df, setup)
        st.success("Đã tạo export.xlsx")
    if os.path.exists(EXPORT_XLSX):
        with open(EXPORT_XLSX, "rb") as f:
            st.download_button("Tải export.xlsx", f, file_name="export.xlsx")
